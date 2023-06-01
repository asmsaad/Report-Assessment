from tkinter import *
from tkinter import messagebox
from tkinter import filedialog
import os,json,sys,pywintypes

import openpyxl
from PIL import Image ,ImageTk,ImageDraw



def center_window(window,width=0,height=0):
    ws = window.winfo_screenwidth()
    hs = window.winfo_screenheight()
    x = (ws/2) - (width/2)
    y = (hs/2) - (height/2)
    window.geometry('%dx%d+%d+%d' % (width, height, x, y))




def progress_image(percentage=1):
    base_width = 400
    base_height = 30
    if float(percentage) > 100: percentage = 100
    percentage = int(percentage)
    percentage = percentage*4
    if percentage<= 35: percentage =35

    base_img = Image.new('RGBA',(base_width,base_height))
    canvas = ImageDraw.Draw(base_img)
    canvas.rounded_rectangle((5,5,base_width-5,base_height-5), radius=10, fill='#DDF2FF', outline='#97B2AF', width=2)
    if percentage> 35:
        canvas.rounded_rectangle((7,7,percentage-7,base_height-7), radius=10, fill='#00A89D', outline='#97B2AF', width=0)    
    return base_img   




class Assasment():
    colorTk = {
        'white' : 'white',
        'gray1' : '#65666A',
        'gray2' : '#F9F9F9',
        'gray3' : '#DFE5EE',
        'blue1' :'#326297',
    }


    # question_dict = {
    #     "1" : "Clarity of language in the report.",
    #     "2" : "Support of the report's findings or impressionss.",
    #     "3" : "Organization and visual clarity of the report.",
    #     "4" : "Effective structuring and presentation of information.",
    #     "5" : "Use of language to convey complex scientific concepts.",
    #     "6" : "Accuracy of scientific information in the report.",
    #     "7" : "Relevance of the report to its intended scientific purpose.",
    #     "8" : "Adherence to scientific standards and norms in tone, style, and presentation.",
    #     "9" : "saad",
    #     "10" : "Mia"
    # }

    answer_dict = {
        "Very Unsatisfied" : "very_unsatisfied",
        "Unsatisfied" : "unsatisfied",
        "Neutral" : "neutral",
        "Satisfied" : "satisfied",
        "Very Satisfied" : "very_satisfied"
    }




    def __init__(self,BASE):
        BASE['bg'] = self.colorTk['white']
        global imageTk
        self.imageTk = imageTk
        self.BASE = BASE
        # Define Window icon and title
        app_icon = Image.open("res/radiation.png")
        app_icon = ImageTk.PhotoImage(app_icon.resize((16,16),Image.LANCZOS))
        self.BASE.iconphoto(False, app_icon)
        self.BASE.title('Report Assessment')
        # self.BASE.geometry('1300x750')
        center_window(self.BASE,width=1300,height=750)
        # On resize action
        self.BASE.bind("<Configure>", self.on_window_resize)

        # Working Frame
        self.working_frame = Frame(self.BASE,bg=self.colorTk['white'],border=0,borderwidth=0,highlightthickness=0)
        #Data load Frame
        self.data_load_frame = Frame(self.BASE,bg=self.colorTk['white'],border=0,borderwidth=0,highlightthickness=0)
        self.data_load_frame.pack(anchor=W)
        self.data_load_frame_property()


    ''' This method is the responsible for task'''
    def working_frame_property(self):
        # Menu Frame
        self.menu_frame = Frame(self.working_frame,bg=self.colorTk['blue1'],border=0,borderwidth=0,highlightthickness=0,pady=6)
        self.menu_frame.pack(anchor=W,expand=True,fill=X)

        # Left Margin
        Label(self.menu_frame,font=('Arial','12','normal'),bg=self.colorTk['blue1'],padx=4).pack(side=LEFT,anchor='nw')
        # Database Icon 
        self.database_BTN= Button(self.menu_frame,bg=self.colorTk['blue1'],activebackground=self.colorTk['blue1'],height=26,width=26,border=0,borderwidth=0,highlightthickness=0)
        self.database_BTN['image'] = imageTk['database']
        self.database_BTN.image = imageTk['database']
        self.database_BTN.pack(anchor='w',side=LEFT)
        # Middle Margin
        Label(self.menu_frame,font=('Arial','1','normal'),bg=self.colorTk['blue1'],padx=0).pack(side=LEFT,anchor='nw')
        # Database Location Entry
        self.database_loc_Entry = Entry(self.menu_frame,font=('Arial','12','normal'),width=len(self.csv_xl_file_location),bg=self.colorTk['blue1'],fg=self.colorTk['gray3'],border=0,borderwidth=0,highlightthickness=0)
        self.database_loc_Entry.pack(anchor='w',side=LEFT)
        self.database_loc_Entry.insert(0,self.csv_xl_file_location)
        self.database_loc_Entry.bind('<KeyRelease>', lambda e,  widget = self.database_loc_Entry , value = self.csv_xl_file_location: self.make_notchange(e,widget,value)) 
        

        # Right Margin
        Label(self.menu_frame,font=('Arial','12','normal'),bg=self.colorTk['blue1'],padx=4).pack(side=RIGHT,anchor='nw')
        #Save Icon
        self.save_BTN= Button(self.menu_frame,bg=self.colorTk['blue1'],activebackground=self.colorTk['blue1'],height=26,width=26,border=0,borderwidth=0,highlightthickness=0)
        self.save_BTN['image'] = imageTk['save']
        self.save_BTN.image = imageTk['save']
        self.save_BTN.pack(anchor='e',side=RIGHT)
        # Save Location Entry
        output_file = os.path.split(self.csv_xl_file_location) 
        self.output_file = "/".join(output_file[:-1])+'/Report_Output.csv'
        self.save_Entry = Entry(self.menu_frame,font=('Arial','12','normal'),width=len(self.output_file),bg=self.colorTk['blue1'],fg=self.colorTk['gray3'],border=0,borderwidth=0,highlightthickness=0)
        self.save_Entry.pack(anchor='e',side=RIGHT)
        self.save_Entry.insert(0,self.output_file)
        self.save_Entry.bind('<KeyRelease>', lambda e,  widget = self.save_Entry ,  value = self.output_file : self.make_notchange(e,widget,value)) 
        
        

        # Top Working Frame
        self.top_working_frame = Frame(self.working_frame,bg=self.colorTk['blue1'],border=0,borderwidth=0,highlightthickness=0,pady=6)
        self.top_working_frame.pack(anchor=W,expand=True,fill=X)

        # Left Margin
        Label(self.top_working_frame,font=('Arial','12','normal'),bg=self.colorTk['blue1'],padx=4).pack(side=LEFT,anchor='nw')
        # Passage Label
        self.section_name_LBL  = Label(self.top_working_frame,text='This is section name',bg=self.colorTk['blue1'],fg=self.colorTk['gray3'],font=('Arial','14','bold'),border=0,borderwidth=0,highlightthickness=0)
        self.section_name_LBL.pack(anchor=W,side=LEFT)
        # Passage Canvas
        self.per_section_identy_Frame  = Canvas(self.top_working_frame,bg=self.colorTk['blue1'],border=0,borderwidth=0,highlightthickness=0)
        self.per_section_identy_Frame.pack(anchor=E,side=RIGHT,expand=True,fill=X)

        Label(self.per_section_identy_Frame,bg=self.colorTk['blue1'],font=('Arial','16','bold'),border=0,borderwidth=0,highlightthickness=0,padx=7).pack(side=RIGHT)

        if self.total_row > 0 :          self.selected_section_index = 1
        else: self.selected_section_index = 0

        self.total_answer = len(list(self.question_dict))*self.total_row
        self.report_assessment_review_dict = {}
        for each_section in range(self.total_row,0,-1):
            # self.report_assessment_review_dict[str(each_section)] = {'1':'','2':'','3':'','4':'','5':'','6':'','7':'','8':'','9':'','10':''}
            self.report_assessment_review_dict[str(each_section)] = {str(i) : '' for i in range(1,len(list(self.question_dict))+1)}
            

        # progress bar
        self.progress_bar = Label(self.per_section_identy_Frame,bg=self.colorTk['blue1'],border=0,borderwidth=0,highlightthickness=0)
        progress_bar_image = ImageTk.PhotoImage(progress_image(0))
        self.progress_bar['image'] = progress_bar_image
        self.progress_bar.image = progress_bar_image
        self.progress_bar.pack(side=RIGHT)

        

        
        # Middle Working Frame
        self.middle_working_frame = Frame(self.working_frame,bg=self.colorTk['white'],border=0,borderwidth=0,highlightthickness=0)
        self.middle_working_frame.pack(anchor=W)



        # Left Margin
        Canvas(self.middle_working_frame,bg=self.colorTk['gray2'],width=12,border=0,borderwidth=0,highlightthickness=0).pack(side=LEFT,anchor='nw',expand=True,fill=Y)
        #Text Widget
        self.text_canvas = Canvas(self.middle_working_frame,bg=self.colorTk['gray2'],border=0,borderwidth=0,highlightthickness=0)
        self.text_canvas.pack(side=LEFT,anchor='nw')

        self.display_text_field = Text(self.text_canvas,font=('Arial','12','normal'),bg=self.colorTk['gray2'],fg=self.colorTk['gray1'],border=0,borderwidth=0,highlightthickness=0)
        # Scroll bar
        scrollbar = Scrollbar(self.text_canvas)
        scrollbar.pack(side=RIGHT, fill=Y)
        # Attach the scrollbar to the text widget
        self.display_text_field.config(yscrollcommand=scrollbar.set)
        scrollbar.config(command=self.display_text_field.yview)
        
        self.display_text_field.pack(anchor=W,expand=True,fill=BOTH,side=RIGHT)



        # Update Passage and Titel
        self.update_display_text_field(self.selected_section_index)
        

        #Right Frame
        # self.question_frame = Frame(self.middle_working_frame,bg=self.colorTk['white'],border=0,borderwidth=0,highlightthickness=0,padx=11)
        # # self.question_frame.pack(side=RIGHT,anchor='ne',expand=True,fill=Y)
        # self.question_frame.pack(side=RIGHT,anchor='ne')

        # create a question_canvas with question_canvas_scrollbar
        self.question_frame_0 = Canvas(self.middle_working_frame,bg=self.colorTk['white'],border=0,borderwidth=0,highlightthickness=0)
        self.question_frame_0.pack(anchor='w')

        

        self.question_canvas = Canvas(self.question_frame_0,bg=self.colorTk['white'],border=0,borderwidth=0,highlightthickness=0)
        question_canvas_scrollbar = Scrollbar(self.question_frame_0, orient="vertical", command=self.question_canvas.yview)
        self.question_frame = Frame(self.question_canvas,bg=self.colorTk['white'],border=0,borderwidth=0,highlightthickness=0)

        self.question_frame.bind("<Configure>", lambda e: self.question_canvas.configure(scrollregion=self.question_canvas.bbox("all")))
        self.question_canvas.create_window((0, 0), window=self.question_frame, anchor="nw")
        self.question_canvas.configure(yscrollcommand=question_canvas_scrollbar.set)

        # pack the question_canvas and question_canvas_scrollbar
        self.question_canvas.pack(side="left",anchor='nw',fill=BOTH,expand=True)
        self.question_canvas.bind_all("<MouseWheel>", self._on_mousewheel)

        question_canvas_scrollbar.pack(side="right", fill="y")


        self.question_frame_1 = Frame(self.question_frame,bg=self.colorTk['white'],border=0,borderwidth=0,highlightthickness=0,padx=11)
        self.question_frame_1.pack(anchor='w')

        

        self.question_frame_2 = Frame(self.middle_working_frame,bg=self.colorTk['white'],border=0,borderwidth=0,highlightthickness=0,padx=11)
        self.question_frame_2.pack(anchor='e')

        # Bottom Working Frame
        self.bottom_working_frame = Frame(self.question_frame_2,bg=self.colorTk['white'],border=0,borderwidth=0,highlightthickness=0)
        self.bottom_working_frame.pack(anchor=SE,side=BOTTOM)

        Label(self.bottom_working_frame,bg=self.colorTk['white'],font=('Arial','16','normal'),border=0,borderwidth=0,highlightthickness=0,padx=5).pack(side=RIGHT,anchor='e')
        
        # Right Shift
        if self.total_row > 1 : right_shift_btn_type = "green"
        else : right_shift_btn_type = "gray"
        self.right_button = Button(self.bottom_working_frame,bg=self.colorTk['white'],activebackground=self.colorTk['white'],width=64,height=64,border=0,borderwidth=0,highlightthickness=0)
        self.right_button['image'] = self.imageTk['arrow']['right'][right_shift_btn_type]
        self.right_button.image = self.imageTk['arrow']['right'][right_shift_btn_type]
        self.right_button['command'] = lambda : self.index_maintain(type='+')
        self.right_button.pack(side=RIGHT,anchor='e')

        Label(self.bottom_working_frame,bg=self.colorTk['white'],font=('Arial','16','normal'),border=0,borderwidth=0,highlightthickness=0,padx=25).pack(side=RIGHT,anchor='e')

        # Left Shift
        self.left_button = Button(self.bottom_working_frame,bg=self.colorTk['white'],activebackground=self.colorTk['white'],relief=SUNKEN,width=64,height=64,border=0,borderwidth=0,highlightthickness=0)
        self.left_button['image'] = self.imageTk['arrow']['left']['gray']
        self.left_button.image = self.imageTk['arrow']['left']['gray']
        # self.left_button['command'] = lambda : self.index_maintain(type='-')
        self.left_button.pack(side=RIGHT,anchor='e')

        self.make_question_display()


    def make_notchange(self,e,widgets,value):
        widgets.delete(0,END)
        widgets.insert(0,value)


    def _on_mousewheel(self,event):
        self.question_canvas.yview_scroll(int(-1*(event.delta/120)), "units")



    '''This method is responsible to display the question's and there options'''
    def make_question_display(self):
        # Every Question radio button image dict
        self.each_question_answer_option_image = {}
        # Every Question radio button lable dict
        self.each_question_answer_option_label = {}

        count_question = 0
        for each_question in self.question_dict:
            count_question += 1

            # Every Question radio button image dict
            self.each_question_answer_option_image[str(count_question)] = {}
            # Every Question radio button lable dict
            self.each_question_answer_option_label[str(count_question)] = {}

            self.each_question_frame = Frame(self.question_frame_1,bg=self.colorTk['white'],border=0,borderwidth=0,highlightthickness=0,pady=3)
            self.each_question_frame.pack(anchor='w',side=TOP)
            
            Label(self.each_question_frame,font=('Arial','14','bold'),bg=self.colorTk['white'],width=60).grid(row=1,column=1,columnspan=2,sticky=W)
            Label(self.each_question_frame,font=('Arial','14','bold'),bg=self.colorTk['white'],width=60).grid(row=2,column=1,columnspan=2,sticky=W)
            
            each_question_frame2 = Frame(self.each_question_frame,bg=self.colorTk['white'],border=0,borderwidth=0,highlightthickness=0,pady=0)
            each_question_frame2.grid(row=1,column=1,columnspan=2,sticky=W)

            
            
            Label(each_question_frame2,text=f'{each_question}.',font=('Arial','14','bold'),bg=self.colorTk['white'],fg=self.colorTk['blue1'],border=0,borderwidth=0,highlightthickness=0,pady=2).grid(row=1,column=1,sticky=W)
            Label(each_question_frame2,text=f'{self.question_dict[each_question]}',font=('Arial','14','normal'),bg=self.colorTk['white'],fg=self.colorTk['blue1'],border=0,borderwidth=0,highlightthickness=0,pady=2).grid(row=1,column=2,sticky=W)
            self.question_options(self.each_question_frame,count_question,each_question)

            Label(self.each_question_frame,bg=self.colorTk['white'],font=('Arial','1','bold'),border=0,borderwidth=0,highlightthickness=0,pady=0).grid(row=3,column=1,columnspan=2)
            Frame(self.each_question_frame,bg='#D9D9D9',border=0,borderwidth=10,highlightthickness=0,height=1,width=700,pady=0).grid(row=4,column=1,columnspan=2,sticky=W)
            Label(self.each_question_frame,bg=self.colorTk['white'],font=('Arial','1','bold'),border=0,borderwidth=0,highlightthickness=0,pady=1).grid(row=5,column=1,columnspan=2)
        

        
        


    def index_maintain(self,type='+'):
        self.question_canvas.yview_moveto(0)
        if type == '+' :
            if self.selected_section_index + 1 <= self.total_row:
                self.selected_section_index += 1
                # Right shift Button Status
                if self.selected_section_index+1 <= self.total_row:
                    self.right_button['image'] = self.imageTk['arrow']['right']['green']
                    self.right_button.image = self.imageTk['arrow']['right']['green']
                    self.right_button['relief'] = RAISED
                    self.right_button['command'] = lambda : self.index_maintain(type='+') 

                    self.left_button['image'] = self.imageTk['arrow']['left']['green']
                    self.left_button.image = self.imageTk['arrow']['left']['green']
                    self.left_button['command'] = lambda : self.index_maintain(type='-')

                else:
                    self.right_button['image'] = self.imageTk['arrow']['right']['gray']
                    self.right_button.image = self.imageTk['arrow']['right']['gray']
                    self.right_button['relief'] = SUNKEN
                    self.right_button['command'] = ''



        elif type == '-' :
            if self.selected_section_index - 1 >= 1:
                self.selected_section_index -= 1
                # Left shift Button Status
                if self.selected_section_index-1 >= 1:
                    self.left_button['image'] = self.imageTk['arrow']['left']['green']
                    self.left_button.image = self.imageTk['arrow']['left']['green']
                    self.left_button['relief'] = RAISED
                    self.left_button['command'] = lambda : self.index_maintain(type='-')

                    self.right_button['image'] = self.imageTk['arrow']['right']['green']
                    self.right_button.image = self.imageTk['arrow']['right']['green']
                    self.right_button['command'] = lambda : self.index_maintain(type='+') 
                else:
                    self.left_button['image'] = self.imageTk['arrow']['left']['gray']
                    self.left_button.image = self.imageTk['arrow']['left']['gray']
                    self.left_button['relief'] = SUNKEN
                    self.left_button['command'] = ''

    

        # Update Passage and Titel
        self.update_display_text_field(self.selected_section_index)
        print('\n\n')
        for question_index in  self.report_assessment_review_dict[str(self.selected_section_index)]:
            print(question_index)
            reaction_index = str(self.report_assessment_review_dict[str(self.selected_section_index)][question_index])
            for reaction_index in range(1,6):
                self.each_question_answer_option_image[str(question_index)][str(reaction_index)]['image'] = self.imageTk['radio']['gray']['16'] 
                self.each_question_answer_option_image[str(question_index)][str(reaction_index)].image = self.imageTk['radio']['gray']['16']
            
        for question_index in  self.report_assessment_review_dict[str(self.selected_section_index)]:
            reaction_index = str(self.report_assessment_review_dict[str(self.selected_section_index)][question_index])
            if reaction_index != '':
                self.each_question_answer_option_image[str(question_index)][str(reaction_index)]['image'] = self.imageTk['radio']['blue']['16'] 
                self.each_question_answer_option_image[str(question_index)][str(reaction_index)].image = self.imageTk['radio']['blue']['16']

        
            



 



    def update_display_text_field(self,row):
        #Change Text Widget
        text_titel =  self.xl_sheet_data.cell(row=row,column=1).value
        text_value =  self.xl_sheet_data.cell(row=row,column=2).value
        self.display_text_field.delete(1.0,END)
        self.display_text_field.insert(1.0,text_value)
        self.section_name_LBL['text'] = text_titel
    

    '''This method help to display the all questions options '''
    def question_options(self,frame,count_question,each_question):
        question_index = count_question

        # Every Question Dictionary 
       

        # Adjustment Label
        Label(frame,bg=self.colorTk['white'],font=('Arial','11','normal'),border=0,borderwidth=0,highlightthickness=0,padx=8).grid(row=2,column=1,sticky=W)
        # Option Frame
        options_frame = Frame(frame,bg=self.colorTk['white'],border=0,borderwidth=0,highlightthickness=0)
        options_frame.grid(row=2,column=2,sticky=W)
        # Create all options
        reaction_index =0    
        for each_answer in self.answer_dict:
            reaction_index += 1

            # Every Question radio button image dict
            self.each_question_answer_option_image[str(question_index)][str(reaction_index)] = Button(options_frame,bg=self.colorTk['white'],activebackground=self.colorTk['white'],height=20,width=20,border=0,borderwidth=0,highlightthickness=0)
            self.each_question_answer_option_image[str(question_index)][str(reaction_index)]['image'] = self.imageTk['radio']['gray']['16'] 
            self.each_question_answer_option_image[str(question_index)][str(reaction_index)].image = self.imageTk['radio']['gray']['16']
            self.each_question_answer_option_image[str(question_index)][str(reaction_index)].pack(side=LEFT,anchor='w')
            self.each_question_answer_option_image[str(question_index)][str(reaction_index)]['command'] = \
                lambda  question_index = question_index , reaction_index = reaction_index : self.select_question_option(question_index,reaction_index)
            
            # Every Question radio button lable dict
            option_LBL_Frame = Frame(options_frame,bg=self.colorTk['white'],border=0,borderwidth=0,highlightthickness=0)
            option_LBL_Frame.pack(side=LEFT,anchor='s')
            Label(option_LBL_Frame,width=13,font = ('Arial','12','normal'),bg=self.colorTk['white'],border=0,borderwidth=0,highlightthickness=0).grid(row=1,column=1,sticky=NW)
            
            self.each_question_answer_option_label[str(question_index)][str(reaction_index)] = Button(option_LBL_Frame,text=each_answer,font = ('Arial','11','normal'),fg=self.colorTk['blue1'],bg=self.colorTk['white'],activebackground=self.colorTk['white'],relief=SUNKEN,border=0,borderwidth=0,highlightthickness=0,padx=2)
            self.each_question_answer_option_label[str(question_index)][str(reaction_index)].grid(row=1,column=1,sticky=NW)
            self.each_question_answer_option_label[str(question_index)][str(reaction_index)]['command'] = \
                lambda  question_index = question_index , reaction_index = reaction_index : self.select_question_option(question_index,reaction_index)

            


    def select_question_option(self,question_index,reaction_index):
        for each_reaction_index in range(1,len(self.answer_dict.keys())+1):
            self.each_question_answer_option_image[str(question_index)][str(each_reaction_index)]['image'] = self.imageTk['radio']['gray']['16'] 
            self.each_question_answer_option_image[str(question_index)][str(each_reaction_index)].image = self.imageTk['radio']['gray']['16']

        self.each_question_answer_option_image[str(question_index)][str(reaction_index)]['image'] = self.imageTk['radio']['blue']['16'] 
        self.each_question_answer_option_image[str(question_index)][str(reaction_index)].image = self.imageTk['radio']['blue']['16']

        # update input to dict
        self.report_assessment_review_dict[str(self.selected_section_index)][str(question_index)] = int(reaction_index)

        # Find Answerd data
        total_given_answer = 0
        for each_question in self.report_assessment_review_dict:
            for each_option in self.report_assessment_review_dict[each_question]:
                if self.report_assessment_review_dict[each_question][each_option] != '':
                    total_given_answer += 1

        if total_given_answer == 0 : total_given_answer = 1

        # Show progressbar
        progress_bar_percentage = (total_given_answer / self.total_answer)*100
        progress_bar_image = ImageTk.PhotoImage(progress_image(progress_bar_percentage))
        self.progress_bar['image'] = progress_bar_image
        self.progress_bar.image = progress_bar_image

        if progress_bar_percentage == 100 and self.show_done_message == True:
            self.show_done_message = False
            messagebox.showinfo(title="Information", message="Excellent work! You have successfully answered all the questions.")

        # Save the data
        self.save_real_time()



    def save_real_time(self):
        try:
            data = 'Passage ID,' + ",".join([f"Q{str(i)}" for i in range(1,len(list(self.question_dict.keys()))+1)])+"\n"
            with open(self.output_file,'w') as RF:
                for each_passage_index in self.report_assessment_review_dict:
                    each_question_line = ''
                    for each_question_index in self.report_assessment_review_dict[each_passage_index]:
                        each_reaction_index = self.report_assessment_review_dict[each_passage_index][each_question_index]
                        each_question_line = each_question_line +str(each_reaction_index) + ' ,'
                    data = data + str(self.xl_sheet_data.cell(row=int(each_passage_index),column=1).value) + ',' + each_question_line+'\n'
                RF.write(f'{data}') 
            RF.close()
        except Exception as e:
            messagebox.showerror("showerror", "Please close the file Report_Output.csv")
            # print(e)
            


    def data_load_frame_property(self):
        #Top margin for align  vartical middle
        self.data_load_top_margin_LBL = Label(self.data_load_frame,bg=self.colorTk['white'],border=0,borderwidth=0,highlightthickness=0)
        self.data_load_top_margin_LBL.pack()
        # Browse Image for CSV or EXCEL file
        self.browse_file_BTN = Button(self.data_load_frame,image = self.imageTk['excel']['gray']['256'],bg=self.colorTk['white'],activebackground=self.colorTk['white'],height=256+5,width=256,border=0,borderwidth=0,highlightthickness=0)
        self.browse_file_BTN.image = self.imageTk['excel']['gray']['256']
        self.browse_file_BTN['command'] = self.browse_excel_csv_file
        self.browse_file_BTN.pack()
        # Label for browsing
        self.browse_LBL = Label(self.data_load_frame,text='Browse Report Assessment',bg=self.colorTk['white'],border=0,borderwidth=0,highlightthickness=0,font=('Arial','16','bold'),fg='#939393',pady=7)
        self.browse_LBL.pack()
        # Loading animation Frame
        self.data_loading_animation_frame = Frame(self.data_load_frame,bg=self.colorTk['white'],border=0,borderwidth=0,highlightthickness=0)
        self.data_loading_animation_frame.pack()





    # BROWSE FAIL
    def browse_excel_csv_file(self):
        self.csv_xl_file_location = filedialog.askopenfilename(initialdir='./',title='Browse input Excel file',filetypes=(('Excel file','*.xlsx'),))
        if self.csv_xl_file_location != None:
            if os.path.isfile(self.csv_xl_file_location) and self.csv_xl_file_location.strip().lower().endswith('.xlsx'):
                self.open_xlsx_file(self.csv_xl_file_location)
                self.browse_LBL['text'] = ''
                self.browse_LBL['fg'] = '#23A566'
                self.browse_file_BTN['image'] = self.imageTk['excel']['green']['256']
                self.browse_file_BTN.image = self.imageTk['excel']['green']['256']
                self.browse_file_BTN['command'] = ''
                self.topic_calulate_animation()

    def open_xlsx_file(self,path):
        work_book = openpyxl.load_workbook(path)
        # self.xl_sheet_data = work_book.active
        self.xl_sheet_data = work_book.worksheets[0]
        self.total_row = self.xl_sheet_data.max_row

        try :
            self.question_sheet = work_book.worksheets[1]
            max_row_question  = self.question_sheet.max_row
        except:
            messagebox.showinfo(title="Error", message="Could not find the question sheet on the spreadsheet.")


        self.show_done_message = True
        
        self.question_dict = {}

        if "question" in str(self.question_sheet.cell(row=1,column=1).value).lower() : start_row = 1
        else: start_row = 2
        for i in range(start_row,max_row_question+1):
            if start_row == 2: i -= 1
            self.question_dict[str(i)] = str(self.question_sheet.cell(row=i,column=1).value)





        
    # EVENT LISTNER
    def on_window_resize(self,event):
        try:
            self.data_load_top_margin_LBL['pady'] = (WINDOW.winfo_height() - self.browse_file_BTN['height']) /6
            self.browse_file_BTN['width'] = self.BASE.winfo_width()
        except: pass

        try:
            # Text Canvas
            self.question_frame_0['width'] = 750
            self.question_frame_0['height'] = self.BASE.winfo_height() - 150
            self.text_canvas['width'] = self.BASE.winfo_width()-800+38
            self.text_canvas['height'] = self.BASE.winfo_height()
            # self.BASE.update()
        except: pass


    # Animation
    def topic_calulate_animation(self,current_index=0,max_index=10,import_dot=1):
        current_index += 1
        duration = int(1200/max_index)
         

        self.browse_LBL['text'] = f'Importing{"."*import_dot}'
        import_dot += 1
        if import_dot >3 : import_dot = 1

        loading_item_ILBL = Label(self.data_loading_animation_frame,image=imageTk['loading1']['green']['32'],bg=self.colorTk['white'],border=0,borderwidth=0,highlightthickness=0)
        loading_item_ILBL.image = imageTk['loading1']['green']['32']
        loading_item_ILBL.pack(side=LEFT)
        
        if current_index<=max_index:
            self.BASE.after(duration,self.topic_calulate_animation,current_index,max_index,import_dot)
        else:
            self.browse_LBL['text'] = 'Imported'
            self.browse_file_BTN['image'] = self.imageTk['excel']['conform']['256']
            self.browse_file_BTN.image = self.imageTk['excel']['conform']['256']
            self.BASE.update()
            self.BASE.after(1500)
            # Disappear
            self.data_load_frame.pack_forget()
            self.working_frame.pack()

            self.working_frame.pack(anchor=W)
            self.working_frame_property()

        




WINDOW = Tk()
global imageTk
imageTk = \
    { "excel" : 
        { "gray" : 
            { "256" :  ImageTk.PhotoImage(Image.open('res/excel_gray.png').resize((256,256),Image.LANCZOS))
        
            },
        "green" :{ "256" :  ImageTk.PhotoImage(Image.open('res/excel_green.png').resize((256,256),Image.LANCZOS))
        
            },
        "conform" : { "256" : ImageTk.PhotoImage(Image.open('res/check_mark.png').resize((256,256),Image.LANCZOS)),
    
            },

        },

    "loading1" :
        { "gray" : 
            { "32" : ''
    
            },
        "green":
            { "32" : ImageTk.PhotoImage(Image.open('res/rectangle_green.png').resize((32,32),Image.LANCZOS))
    
            }
        },
    "radio" :
        { "gray" : 
            { "16" : ImageTk.PhotoImage(Image.open('res/radio_gray.png').resize((16,16),Image.LANCZOS))
    
            },
        "blue":
            { "16" : ImageTk.PhotoImage(Image.open('res/radio.png').resize((16,16),Image.LANCZOS))
    
            }
    
        },
    "arrow" :  
        { "left" : 
            { "gray" : ImageTk.PhotoImage(Image.open('res/left_arrow_gray.png').resize((60,60),Image.LANCZOS)),
            "green" : ImageTk.PhotoImage(Image.open('res/left_arrow.png').resize((60,60),Image.LANCZOS))         
            },
        "right" : 
            { "gray" : ImageTk.PhotoImage(Image.open('res/right_arrow_gray.png').resize((60,60),Image.LANCZOS)),
            "green" : ImageTk.PhotoImage(Image.open('res/right_arrow.png').resize((60,60),Image.LANCZOS))         
            },
        },

    
    "database" : ImageTk.PhotoImage(Image.open('res/database.png').resize((24,24),Image.LANCZOS)),
    "save" : ImageTk.PhotoImage(Image.open('res/save.png').resize((20,20),Image.LANCZOS)),
   

    }
Assasment(WINDOW)
WINDOW.mainloop()